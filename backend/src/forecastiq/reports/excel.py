from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from forecastiq.reports.content import ReportContent, TableSection, TextSection

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_TITLE_FONT = Font(bold=True, name="Calibri", size=16)
_SECTION_FONT = Font(bold=True, name="Calibri", size=12, color="1D4ED8")


def render_excel(content: ReportContent) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = content.title[:31]  # Excel sheet name limit

    sheet.append([content.title])
    sheet["A1"].font = _TITLE_FONT
    sheet.append([f"Generated at {content.generated_at}"])
    sheet.append([])

    for section in content.sections:
        sheet.append([section.title])
        sheet.cell(row=sheet.max_row, column=1).font = _SECTION_FONT
        if isinstance(section, TextSection):
            for line in section.body.split("\n"):
                sheet.append([line])
                sheet.cell(row=sheet.max_row, column=1).alignment = Alignment(wrap_text=True)
        elif isinstance(section, TableSection):
            header_row = sheet.max_row + 1
            sheet.append(section.columns)
            for col_idx in range(1, len(section.columns) + 1):
                cell = sheet.cell(row=header_row, column=col_idx)
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            for row in section.rows:
                sheet.append(list(row))
                for col_idx in range(1, len(section.columns) + 1):
                    sheet.cell(row=sheet.max_row, column=col_idx).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )
        sheet.append([])

    # Auto-fit-ish column widths so long messages aren't clipped in Excel.
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 72))
        sheet.column_dimensions[letter].width = max(12, max_len + 2)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
